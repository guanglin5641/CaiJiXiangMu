import openpyxl
from googletrans import Translator
def translate_text(text, target_lang):
    translator = Translator()
    translation = translator.translate(text, dest=target_lang).text
    return translation
def translate_excel(file_path, target_lang):
    wb = openpyxl.load_workbook(file_path)
    translated_wb = openpyxl.Workbook()
    translated_sheet = translated_wb.active
    for sheet in wb.sheetnames:
        current_sheet = wb[sheet]
        translated_sheet = translated_wb.create_sheet(title=sheet)
        for row in current_sheet.iter_rows():
            translated_row = []
            for cell in row:
                cell_value = cell.value
                if cell_value:
                    translated_text = translate_text(cell_value, target_lang)
                    translated_row.append(translated_text)
                else:
                    translated_row.append(None)
            translated_sheet.append(translated_row)
    translated_wb.save('translated_' + file_path)
 # 示例用法
file_path = r'C:\Users\86176\Desktop\Amazon/aaa.xlsx'
target_lang = 'en'  # 目标语言，例如英语
translate_excel(file_path, target_lang)