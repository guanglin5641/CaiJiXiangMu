from openpyxl import load_workbook
from openpyxl.drawing.image import Image
 # 打开Excel文件
workbook = load_workbook(r'C:\Users\86176\Desktop\Amazon/111.xlsx')
 # 获取目标工作表
worksheet = workbook['your_worksheet']
 # 遍历工作表的所有行
for row in worksheet.iter_rows():
    for cell in row:
        # 检查单元格是否包含图片链接
        if cell.hyperlink and cell.hyperlink.is_external:
            # 获取图片文件路径
            image_file_path = cell.hyperlink.target
            # 修改图片链接
            new_image_file_path = image_file_path.replace('.jpg', '.jpeg')
            # 更新单元格的值为修改后的链接
            cell.hyperlink.target = new_image_file_path
 # 保存修改后的Excel文件
workbook.save('modified_excel_file.xlsx')

def test():
    a = "Autumn And Winter Women'S Knitwear Cheongsam Collar Women'S V-Neck Long-Sleeved Bottoming Shirt Women'S Round Neck Sweater Women'S 100% Pure Cashmere Sweater,Rose Purple,3XL"
    b = "Rose Purple"
    c = "3XL"
    d = a.replace(b, "",)
    e = d.replace(c, "",)
    print(e)
    return e
if __name__ == '__main__':
    test()
